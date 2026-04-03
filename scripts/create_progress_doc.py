# -*- coding: utf-8 -*-
"""진행사항 문서 생성 스크립트"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

def add_sheet(name, rows):
    ws = wb.create_sheet(name)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
    return ws

# 시트 1: 진행 요약
summary = [
    ["양주 정원 서비스 개발 진행사항", "", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["작업일", "2025-03-10", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["완료 Phase", "Phase 1 (지도 MVP)", "", "", "", "", ""],
    ["구현 범위", "SVG 지도, 섹션 하이라이트, 회전 버튼, 섹션 팝오버, 섹션 드로어, 헤더, 탭(지도/할 일), 할 일 페이지 기본 UI", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["다음 단계", "Phase 2: 할 일 CRUD, 필터/정렬/그룹화", "Phase 3: 식물 DB", "Phase 4: 공유", "", "", ""],
]
add_sheet("진행요약", summary)

# 시트 2: 구현 완료 항목
completed = [
    ["구분", "ID", "항목", "상태", "비고"],
    ["페이지", "PG-01", "랜딩 페이지(지도)", "완료", "GardenMap 컴포넌트"],
    ["페이지", "PG-02", "할 일 전체 페이지", "완료", "TaskListView 기본 UI"],
    ["페이지", "PG-09", "공통 헤더/네비", "완료", "Header 컴포넌트"],
    ["컴포넌트", "CP-01", "로고", "완료", "Header 내"],
    ["컴포넌트", "CP-02", "탭(지도/할 일)", "완료", "Header 내"],
    ["컴포넌트", "CP-04", "SVG 간이 지도", "완료", "GardenMap.jsx"],
    ["컴포넌트", "CP-05", "회전 버튼", "완료", "도로/집 기준 전환"],
    ["컴포넌트", "CP-06", "섹션 하이라이터", "완료", "레전드 + SVG 클릭"],
    ["컴포넌트", "CP-08", "섹션 팝오버", "완료", "hover 시 표시"],
    ["컴포넌트", "CP-10", "공통 드로어 쉘", "완료", "SectionDrawer"],
    ["컴포넌트", "CP-11", "섹션 디테일", "완료", "금주 할 일, 식물 수"],
    ["기능", "FN-01", "지도 회전", "완료", ""],
    ["기능", "FN-02", "섹션 하이라이트(필터)", "완료", ""],
    ["기능", "FN-03", "섹션 하이라이트(지도)", "완료", ""],
    ["기능", "FN-04", "섹션 팝오버", "완료", ""],
    ["기능", "FN-05", "섹션 드로어 오픈", "완료", ""],
    ["기능", "FN-07", "금주 할 일 목록 노출", "완료", ""],
    ["기능", "FN-09", "할 일 기본 조회", "완료", "섹션 그룹, 예정일 순"],
    ["기능", "FN-12", "그룹화", "완료", "섹션별 그룹"],
    ["기능", "FN-22", "네비게이션 상태 유지", "완료", "React Router"],
]
add_sheet("구현완료", completed)

# 시트 3: 미구현/예정
pending = [
    ["Phase", "ID", "항목", "우선순위", "비고"],
    ["2", "FN-10", "property 기준 필터", "P1", "할 일 필터"],
    ["2", "FN-11", "property 기준 정렬", "P1", ""],
    ["2", "FN-13", "할 일 생성", "P2", ""],
    ["2", "FN-14", "수정/완료/삭제", "P1", "CRUD 액션"],
    ["3", "PG-05", "식물 드로어", "P2", ""],
    ["3", "PG-07", "식물 전체 페이지", "P2", ""],
    ["3", "FN-15~19", "식물 관련 기능", "P2", ""],
    ["4", "PG-03", "공유 페이지", "P4", ""],
    ["4", "FN-20", "공유 URL 발행", "P4", ""],
]
add_sheet("미구현예정", pending)

# 시트 4: 기술 스택
tech = [
    ["구분", "항목"],
    ["프레임워크", "Vite + React 18"],
    ["라우팅", "react-router-dom v6"],
    ["데이터", "임시 mockData (API 연동 대비)"],
    ["스타일", "CSS (CSS 변수 활용)"],
    ["배포", "Vercel 권장 (DEPLOYMENT.md 참고)"],
]
add_sheet("기술스택", tech)

wb.save("doc/진행사항.xlsx")
print("doc/진행사항.xlsx 생성 완료")
